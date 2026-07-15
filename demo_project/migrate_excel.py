"""One-time import: web_links.xlsx -> library.db. Reads the workbook READ-ONLY.

Duplicates (same desc_key across hq/grave/new, now site-agnostic) are MERGED:
  - status: 94 or 9 wins; else the highest of {1,2,3,4} (88 ranks just under 4)
  - watch_count: summed        - date_added: oldest        - date_watched: most recent
  - favorite if any copy has keep="T"; stars are unioned.
"""
import os
import sys
from collections import defaultdict
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from db import connect, init_db, DB_PATH, APP_DIR
from identity import canonicalize_url, slugify_link, desc_key, extract_site, split_stars
import scoring
import backup

WORKBOOK = os.path.normpath(os.path.join(APP_DIR, '..', 'web_links.xlsx'))
SHEETS = ['hq', 'grave', 'new']
STATUS_PRIO = {94: 100, 9: 90, 4: 80, 88: 75, 3: 70, 2: 60, 1: 50, None: 0}


def _iso(v):
    return v.strftime('%Y-%m-%d') if isinstance(v, (datetime, date)) else None


def _dt(v):
    return v if isinstance(v, (datetime, date)) else None


def _rows(ws):
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    for r in it:
        if not all(x is None for x in r):
            yield dict(zip(hdr, r))


def _num(v):
    return int(v) if isinstance(v, (int, float)) else None


def merge_group(members):
    primary = max(members, key=lambda m: STATUS_PRIO.get(m['status'], 0))
    status = primary['status']
    processed = any(m['sheet'] in ('hq', 'grave') for m in members)
    if status == 4 or (status is None and processed):
        location = 'watch'
    elif status is None and not processed:
        location = 'new'
    else:
        location = 'grave'
    added = [_dt(m['date_added']) for m in members if _dt(m['date_added'])]
    watched = [_dt(m['date_watched']) for m in members if _dt(m['date_watched'])]
    stars = []
    for m in members:
        for s in m['stars']:
            if s not in stars:
                stars.append(s)
    tags = primary['tags'] or max((m['tags'] for m in members), key=lambda t: len(t or ''), default=None)
    return {
        'desc_key': primary['desc_key'], 'link_desc': primary['link_desc'],
        'url': primary['url'], 'site': primary['site'], 'location': location,
        'status': status, 'favorite': 1 if any(m['favorite'] for m in members) else 0,
        'watch_count': sum(m['watch_count'] for m in members),
        'date_added': min(added).strftime('%Y-%m-%d') if added else None,
        'date_watched': max(watched).strftime('%Y-%m-%d') if watched else None,
        'tags': tags, 'stars': stars,
    }


def migrate(workbook=WORKBOOK, db_path=DB_PATH):
    if os.path.exists(db_path):
        backup.snapshot('pre-migrate')
        os.remove(db_path)
    conn = connect(db_path)
    init_db(conn)

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    groups = defaultdict(list)
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        for r in _rows(wb[sheet]):
            link = r.get('Link')
            if not link:
                continue
            url = canonicalize_url(str(link))
            ld = r.get('link_desc') or slugify_link(url)
            key = desc_key(ld)
            if not key:
                continue
            stars = split_stars(r.get('Stars'))
            groups[key].append({
                'sheet': sheet, 'desc_key': key, 'link_desc': ld, 'url': url,
                'site': extract_site(url), 'status': _num(r.get('Status')),
                'watch_count': _num(r.get('WatchedCount')) or 0,
                'favorite': 1 if str(r.get('keep') or '').strip() == 'T' else 0,
                'date_added': r.get('DateAdded'), 'date_watched': r.get('Date'),
                'tags': (str(r.get('tags')) if r.get('tags') not in (None, '', '0', 'NA', '#N/A') else None),
                'stars': stars,
            })

    actor_id = {}

    def get_actor(name):
        if name not in actor_id:
            conn.execute("INSERT OR IGNORE INTO actors(name) VALUES (?)", (name,))
            actor_id[name] = conn.execute("SELECT id FROM actors WHERE name=?", (name,)).fetchone()['id']
        return actor_id[name]

    n_groups = n_merged = 0
    for key, members in groups.items():
        n_groups += 1
        if len(members) > 1:
            n_merged += 1
        m = merge_group(members)
        cur = conn.execute("""
            INSERT INTO videos (desc_key, link_desc, url, site, location, status, favorite,
                watch_count, date_added, date_watched, tags_raw, description_raw)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (m['desc_key'], m['link_desc'], m['url'], m['site'], m['location'], m['status'],
             m['favorite'], m['watch_count'], m['date_added'], m['date_watched'],
             m['tags'], m['link_desc']))
        vid = cur.lastrowid
        for pos, nm in enumerate(m['stars']):
            conn.execute("INSERT OR IGNORE INTO video_actors(video_id,actor_id,position) VALUES (?,?,?)",
                         (vid, get_actor(nm), pos))
    conn.commit()

    stats = scoring.recompute(conn)
    counts = {row['location']: row['n'] for row in conn.execute(
        "SELECT location, COUNT(*) n FROM videos GROUP BY location")}
    backup.snapshot('post-migrate')
    conn.close()
    print(f"groups={n_groups}  merged (>1 row)={n_merged}")
    print(f"by location: {counts}")
    print(f"scoring: {stats}")


if __name__ == '__main__':
    migrate(sys.argv[1] if len(sys.argv) > 1 else WORKBOOK)
